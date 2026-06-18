#!/usr/bin/env python3
"""
Emirates Pride Perfumes — Tester Consumption Report Builder  (v2 — AUDITED)
Periods: Jan 2026 – May 2026 (monthly) + Q1 2026 (quarterly)
Brands : EPP (separate file) + ASL (separate file)

CORRECTED EQUATION (audited 17 Jun 2026):
  T (Testers Dispatched) = tester_history.testers_dispatched
        · SKU normalized: strip trailing -T / -TD / -TN / -TG / -TS  (merge bare + suffixed forms)
        · PAPER CARDS EXCLUDED: EPT*, EPTP*, ASLT*, ASLTP*, names with 'TESTER CARD'/'TESTER PAPER'
  S (Sales Generated)    = sales_history.qty_sold for the bare FG code  (NOT tester_history.sales_generated)
  Contrib% = T ÷ S
  → ONE row per Finished-Good SKU.  e.g. B00015 (Hidden Leather):  T=testers,  S=FG sales.
  → ACCESSORIES sorted to the very bottom of every sheet.

Data sources (pre-aggregated, store kiosk variants collapsed in SQL):
  /tmp/tester_norm.csv : norm_sku, store_code, division, month_year, testers
  /tmp/sales_fg.csv    : sku_code, store_code, month_year, qty_sold
  Category/Sub-Category: SAP replenishment file (product_master is EMPTY in Supabase)
"""

import pandas as pd
import numpy as np
import re, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.expanduser('~/Documents')

# ─── STORE GROUPS (by Area Manager) ──────────────────────────────────────────
# Tester source = warehouse replenishment dispatch. HO (0001) + FNF WH (STR02) are
# non-retail staging destinations but ARE part of total dispatch, shown as a trailing
# group so Grand Total Testers reconciles to the warehouse dispatch totals.
EPP_STORE_GROUPS = [
    ('Mohamed Hessin', [
        ('A0002','BAS-Kiosk'), ('A0001','Bas Shop'), ('A0010','Bas Shop 2'),
        ('A0004','Dalma Kiosk'), ('A0003','Dalma Shop'), ('A0005','Deerfield'),
        ('A0007','Yas Kiosk 2'), ('A0008','Yas Kiosk 3'), ('A0009','Yas Podium'),
        ('A0011','Marina Mall'),
        ('AL001','Al Ain Mall'), ('AL002','Bawadi 1'), ('AL003','Bawadi 2'),
        ('AL004','Jimi Mall'), ('AL006','Makani Shop'),
    ]),
    ('Mohammed Imad', [
        ('DX001','Dubai Mall'), ('DX004','Mall of Emirates'), ('DX005','Mirdif'),
        ('DX006','Dubai Hills'), ('DX008','Festival City'),
    ]),
    ('Mohammed Elmatloub', [
        ('RK001','Manar Shop'), ('RK002','Manar Kiosk'), ('FJ001','Fujairah CC'),
        ('SH001','Zahia CC'), ('AJ001','Ajman CC'),
    ]),
]
ASL_STORE_GROUPS = [
    ('ASL Franchise', [
        ('BAS001','BAS Mall'), ('YMK001','Yas Mall'), ('BAW001','Bawadi Mall'),
        ('MAK001','Makani Mall'), ('FJ0001','Fujairah CC'),
    ]),
]

# ─── COLOURS ─────────────────────────────────────────────────────────────────
EPP_HDR_BG, EPP_HDR2_BG = 'FF1F3864', 'FF283593'
ASL_HDR_BG, ASL_HDR2_BG = 'FF006064', 'FF00838F'
WHITE       = 'FFFFFFFF'
LIGHT_BLUE  = 'FFD6E4F7'   # Tester cells
LIGHT_GREEN = 'FFD7F2D7'   # Sales cells
LIGHT_YELLOW= 'FFFFF2CC'   # Contrib% cells
GOLD        = 'FFFFCC00'   # Top performers
ALT_ROW     = 'FFF7F7F7'
TOTALS_ROW  = 'FFEEEEEE'
ACC_BG      = 'FFEDE7DD'   # Accessory section tint
DARK_TEXT   = 'FF1A1A1A'
BLUE_TEXT   = 'FF0D47A1'
GREEN_TEXT  = 'FF1B5E20'
SKU_TEXT    = 'FF1A237E'

def fill(c): return PatternFill('solid', fgColor=c)
def font(b=False,s=10,c='FF000000',i=False): return Font(name='Montserrat',bold=b,size=s,color=c,italic=i)
def align(h='center',v='center',w=True): return Alignment(horizontal=h,vertical=v,wrap_text=w)
THIN=Side(style='thin',color='FFD3D3D3')
def border(): return Border(left=THIN,right=THIN,top=THIN,bottom=THIN)

# ─── ACCESSORY / NON-FRAGRANCE CATEGORIES (sorted to bottom) ─────────────────
ACC_CATS = {'Accessory','Accessories','Bags','Carton','Display','Display - Box',
            'Display - Bottle','Candle','General','Miscellaneous','Offers','Box'}
def is_accessory(sku, cat):
    if str(sku).upper().startswith('AC'): return True
    return str(cat) in ACC_CATS

# ─── CATEGORY + SUB-CATEGORY MAP — AUTHORITATIVE from product_master (610 rows) ─
#     Falls back to SAP replenishment for any SKU not present in the master.
def load_master_map():
    import urllib.request, json
    URL=('https://ncszurcrkngjcjqsowln.supabase.co/rest/v1/product_master'
         '?select=product_code,product_name_en,product_family,category,sub_category')
    KEY=('eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5jc3p1cmNya25n'
         'amNqcXNvd2xuIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzc0NjA4NTgsImV4cCI6MjA5MzAzNjg1OH0.'
         'i5cPlP7JTTCKMXuFqI81WXbjQa71qBkRBZEBvNf6ZmM')
    try:
        req=urllib.request.Request(URL, headers={'apikey':KEY,'Authorization':f'Bearer {KEY}',
            'Range':'0-999'})
        with urllib.request.urlopen(req, timeout=20) as r:
            data=json.loads(r.read())
        out={}
        for p in data:
            code=str(p['product_code']).strip()
            fam=(p.get('product_family') or '').strip()
            cat=(p.get('category') or '').strip()
            sub=(p.get('sub_category') or '').strip()
            out[code]={'Category':fam or cat,
                       'SubCategory':sub or cat,
                       'Name':(p.get('product_name_en') or code).strip()}
        print(f'    · product_master loaded: {len(out)} SKUs (authoritative)')
        return out
    except Exception as e:
        print(f'    · WARN product_master fetch failed ({e}); using SAP fallback only')
        return {}

def load_category_map():
    master = load_master_map()
    rep = pd.read_csv(os.path.join(OUT_DIR,'Replenishment_Jan_May_2026_FINAL.csv'))
    fg = rep[rep['Item_Type']=='FG'].copy()

    def clean_name(name, sku):
        n = re.sub(r'^'+re.escape(str(sku))+r'(-T[A-Z]?)?\s*-\s*','',str(name))
        n = re.sub(r'\s*-Tester.*$','',n,flags=re.I)
        n = re.sub(r'\s+\d+ML\s*$','',n,flags=re.I)
        return n.strip()

    def sub_cat(name, cat):
        s=str(name).lower(); c=str(cat).lower()
        vol = ''
        m = re.search(r'(\d+\s?ml|\d+\s?grm|\d+\s?g\b)', s)
        if m: vol = m.group(1).replace(' ','')
        if 'oil' in s or 'oil' in c:            return f'Oil {vol}'.strip()
        if 'hair' in s or 'mist' in s:          return 'Hair & Body Mist'
        if 'lotion' in s:                       return 'Body Lotion'
        if 'body wash' in c or 'wash' in s:     return 'Body Wash'
        if 'diffuser' in s or 'diffuser' in c:  return 'Diffuser'
        if 'air fresh' in s or 'fresh' in c:    return 'Air Freshener'
        if 'candle' in s or 'candle' in c:      return 'Candle'
        if 'dakhoon' in s or 'dakhoon' in c or 'mukhallat' in s: return 'Dakhoon / Bakhoor'
        if 'spray' in s:                        return 'All Over Spray'
        if any(k in s for k in ['gift','set box','box','bundle','combo','discovery']) \
           or any(k in c for k in ['box','set','gift','offers']): return 'Gift Set / Box'
        if str(sku_placeholder).startswith('AC') if False else False: return 'Accessory'
        if 'tester' in s:                       return 'Tester'
        return f'Perfume {vol}'.strip() if vol else 'Perfume'

    out={}
    for sku, grp in fg.groupby('SKU_Code'):
        cat = grp['Category'].mode().iloc[0] if len(grp['Category'].mode()) else grp['Category'].iloc[0]
        raw = grp['Product_Name'].iloc[0]
        global sku_placeholder; sku_placeholder=sku
        nm  = clean_name(raw, sku)
        sc  = 'Accessory' if str(sku).upper().startswith('AC') else sub_cat(nm, cat)
        out[sku] = {'Category':cat, 'SubCategory':sc, 'Name':nm if nm and nm!=sku else str(sku)}
    # AUTHORITATIVE OVERRIDE: product_master wins where the SKU exists there.
    # Tester codes carry the bare FG code after normalization, so match directly.
    for code, info in master.items():
        out[code] = {'Category':info['Category'] or out.get(code,{}).get('Category',''),
                     'SubCategory':info['SubCategory'] or out.get(code,{}).get('SubCategory',''),
                     'Name':info['Name'] or out.get(code,{}).get('Name',code)}
    return out

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def H(ws,r,c,v,bg,fg='FFFFFFFF',b=True,s=10,h='center',w=True):
    cell=ws.cell(r,c,v); cell.fill=fill(bg); cell.font=font(b,s,fg)
    cell.alignment=align(h,'center',w); cell.border=border(); return cell

def D(ws,r,c,v,bg=None,fg=DARK_TEXT,b=False,s=9,h='center',nf=None):
    cell=ws.cell(r,c,v)
    if bg: cell.fill=fill(bg)
    cell.font=font(b,s,fg); cell.alignment=align(h,'center',True); cell.border=border()
    if nf: cell.number_format=nf
    return cell

# ─── MATRIX SHEET ────────────────────────────────────────────────────────────
# Column layout (Category + Sub-Category added):
# A=#, B=Category, C=Sub-Category, D=SKU Code, E=Product Name,
# F=GrandTotalTesters, G=GrandTotalSales, H=GrandContrib%  then 3 cols/store
FIXED = 8
def build_matrix(ws, brand, tdf, sdf, store_groups, catmap):
    HDR=EPP_HDR_BG if brand=='EPP' else ASL_HDR_BG
    HDR2=EPP_HDR2_BG if brand=='EPP' else ASL_HDR2_BG
    FULL='EMIRATES PRIDE PERFUMES' if brand=='EPP' else 'AROMATIC SCENTS LAB (ASL)'

    flat=[]
    for am,stores in store_groups:
        for code,name in stores: flat.append((am,code,name))
    nS=len(flat)
    def scol(i): return FIXED+1+i*3
    tot_cols=FIXED+nS*3

    # tester lookup  {(sku,store):testers}, sales lookup {(sku,store):qty}
    tl={};
    for _,r in tdf.iterrows(): tl[(r['norm_sku'],r['store_code'])]=tl.get((r['norm_sku'],r['store_code']),0)+int(r['testers'])
    sl={}
    for _,r in sdf.iterrows(): sl[(r['sku_code'],r['store_code'])]=sl.get((r['sku_code'],r['store_code']),0)+int(r['qty_sold'])

    # rows = SKUs that have testers dispatched in this period
    skus=sorted(tdf['norm_sku'].unique())
    store_codes=[c for _,c,_ in flat]

    # build sortable rows: (is_acc, category, sku) with grand totals over displayed stores
    rows=[]
    for sku in skus:
        gt=sum(tl.get((sku,c),0) for c in store_codes)
        gs=sum(sl.get((sku,c),0) for c in store_codes)
        if gt==0 and gs==0: continue
        info=catmap.get(sku,{'Category':'','SubCategory':'','Name':sku})
        acc=is_accessory(sku, info['Category'])
        rows.append({'sku':sku,'cat':info['Category'],'sub':info['SubCategory'],
                     'name':info['Name'] or sku,'gt':gt,'gs':gs,'acc':acc})
    # sort: products first (by cat, sku), accessories last (by cat, sku)
    rows.sort(key=lambda x:(x['acc'], x['cat'] or 'ZZ', x['sku']))

    nR=len(rows)
    R0=9; RL=R0+nR-1 if nR else R0

    # Row 2 title
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=tot_cols)
    c=ws.cell(2,1,f'{FULL} — SKU × STORE TESTER CONSUMPTION REPORT')
    c.fill=fill(HDR); c.font=Font(name='Montserrat',bold=True,size=13,color='FFFFFFFF'); c.alignment=align(w=False)
    ws.row_dimensions[2].height=28
    # Row 3 legend
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=tot_cols)
    c=ws.cell(3,1,'T=Testers Dispatched FG-WH→Store (blue) | S=FG Sales sales_history (green) | Contrib%=T÷S   ·   HO/Warehouse & paper cards excluded · Accessories at bottom')
    c.fill=fill('FFE8EAF6' if brand=='EPP' else 'FFE0F7FA'); c.font=Font(name='Montserrat',size=9,color='FF424242'); c.alignment=align(w=False)
    ws.row_dimensions[3].height=18
    # Row 4 headers
    for ci,h in enumerate(['#','Category','Sub-Category','SKU Code','Product Name',
                            'Grand Total\nTesters','Grand Total\nSales','Grand\nContrib%'],1):
        H(ws,4,ci,h,HDR,s=9)
    ws.row_dimensions[4].height=30
    # Row 5 AM groups
    cur=FIXED+1
    for am,stores in store_groups:
        span=len(stores)*3
        if span>0:
            ws.merge_cells(start_row=5,start_column=cur,end_row=5,end_column=cur+span-1)
            H(ws,5,cur,f'Area Manager: {am}',HDR2,s=9)
        cur+=span
    ws.row_dimensions[5].height=20
    # Row 6 store names
    for i,(am,code,name) in enumerate(flat):
        cs=scol(i); ws.merge_cells(start_row=6,start_column=cs,end_row=6,end_column=cs+2)
        H(ws,6,cs,name,'FF37474F',s=8)
    ws.row_dimensions[6].height=20
    # Row 7 T|S|Contrib%
    for i in range(nS):
        cs=scol(i)
        H(ws,7,cs,'T',LIGHT_BLUE,fg=BLUE_TEXT,s=8)
        H(ws,7,cs+1,'S',LIGHT_GREEN,fg=GREEN_TEXT,s=8)
        H(ws,7,cs+2,'C%',LIGHT_YELLOW,fg='FF5D4037',s=8)
    ws.row_dimensions[7].height=16
    # Row 8 totals
    D(ws,8,1,'TOTAL',bg=TOTALS_ROW,b=True)
    for col in (2,3,4): D(ws,8,col,'',bg=TOTALS_ROW)
    D(ws,8,5,'',bg=TOTALS_ROW,h='left')
    eL,fL=get_column_letter(6),get_column_letter(7)
    if nR:
        ws.cell(8,6,f'=SUM({eL}{R0}:{eL}{RL})'); ws.cell(8,7,f'=SUM({fL}{R0}:{fL}{RL})')
        ws.cell(8,8,f'=IFERROR({eL}8/{fL}8,0)')
    else:
        ws.cell(8,6,0); ws.cell(8,7,0); ws.cell(8,8,0)
    for col,bg in ((6,TOTALS_ROW),(7,TOTALS_ROW),(8,TOTALS_ROW)):
        cc=ws.cell(8,col); cc.fill=fill(bg); cc.font=font(True,10 if col<8 else 9); cc.alignment=align(); cc.border=border()
    ws.cell(8,8).number_format='0.0%'
    for i in range(nS):
        cs=scol(i); tL=get_column_letter(cs); sL=get_column_letter(cs+1)
        if nR:
            ws.cell(8,cs,f'=SUM({tL}{R0}:{tL}{RL})'); ws.cell(8,cs+1,f'=SUM({sL}{R0}:{sL}{RL})')
            ws.cell(8,cs+2,f'=IFERROR({tL}8/{sL}8,"-")')
        else:
            ws.cell(8,cs,0); ws.cell(8,cs+1,0); ws.cell(8,cs+2,'-')
        ws.cell(8,cs).fill=fill(LIGHT_BLUE); ws.cell(8,cs+1).fill=fill(LIGHT_GREEN); ws.cell(8,cs+2).fill=fill(LIGHT_YELLOW)
        for cc in (cs,cs+1,cs+2):
            x=ws.cell(8,cc); x.font=font(True,8); x.alignment=align(); x.border=border()
        ws.cell(8,cs+2).number_format='0.0%'
    ws.row_dimensions[8].height=16

    # data rows
    acc_started=False
    ri=0
    for row in rows:
        r=R0+ri; ri+=1
        if row['acc'] and not acc_started:
            acc_started=True  # (could add a divider; keep simple, tint accessory rows)
        base_bg = ACC_BG if row['acc'] else (WHITE if ri%2 else ALT_ROW)
        D(ws,r,1,ri,bg=base_bg,s=8)
        D(ws,r,2,row['cat'],bg=base_bg,h='left',s=8)
        D(ws,r,3,row['sub'],bg=base_bg,h='left',s=8)
        D(ws,r,4,row['sku'],bg=base_bg,fg=SKU_TEXT,s=9,b=True)
        D(ws,r,5,row['name'],bg=base_bg,h='left',s=9)
        D(ws,r,6,int(row['gt']),bg=LIGHT_BLUE,b=True,s=9)
        D(ws,r,7,int(row['gs']),bg=LIGHT_GREEN,b=True,s=9)
        eR=get_column_letter(6)+str(r); fR=get_column_letter(7)+str(r)
        c8=ws.cell(r,8,f'=IFERROR(IF({fR}>0,{eR}/{fR},"-"),"-")')
        c8.fill=fill(LIGHT_YELLOW); c8.font=font(s=9); c8.alignment=align(); c8.border=border(); c8.number_format='0.0%'
        for i,(am,code,name) in enumerate(flat):
            cs=scol(i); t=tl.get((row['sku'],code),0); s=sl.get((row['sku'],code),0)
            D(ws,r,cs,t if t else '',bg=LIGHT_BLUE,s=9)
            D(ws,r,cs+1,s if s else '',bg=LIGHT_GREEN,s=9)
            tR=get_column_letter(cs)+str(r); sR=get_column_letter(cs+1)+str(r)
            cc=ws.cell(r,cs+2,f'=IFERROR(IF({sR}>0,{tR}/{sR},"-"),"-")')
            cc.fill=fill(LIGHT_YELLOW); cc.font=font(s=8); cc.alignment=align(); cc.border=border(); cc.number_format='0.0%'
        ws.row_dimensions[r].height=15

    # widths
    for col,w in zip('ABCDEFGH',[4,20,18,11,30,10,10,9]):
        ws.column_dimensions[col].width=w
    for i in range(nS):
        cs=scol(i)
        ws.column_dimensions[get_column_letter(cs)].width=6
        ws.column_dimensions[get_column_letter(cs+1)].width=6
        ws.column_dimensions[get_column_letter(cs+2)].width=7
    ws.freeze_panes=ws.cell(9,6)
    return nR

# ─── TOP 10 SHEET ────────────────────────────────────────────────────────────
def build_top10(ws, brand, tdf, sdf, store_groups, catmap):
    HDR=EPP_HDR_BG if brand=='EPP' else ASL_HDR_BG
    FULL='EMIRATES PRIDE PERFUMES' if brand=='EPP' else 'AROMATIC SCENTS LAB (ASL)'
    store_codes=[c for _,sts in store_groups for _,c,_ in [(None,s[0],s[1]) for s in sts]]

    tg=tdf.groupby('norm_sku')['testers'].sum()
    rows=[]
    for sku,t in tg.items():
        if t<=0: continue
        info=catmap.get(sku,{'Category':'','SubCategory':'','Name':sku})
        if is_accessory(sku, info['Category']): continue  # exclude accessories from top performers
        s=int(sdf[sdf.sku_code==sku]['qty_sold'].sum())
        rows.append((sku,info['Category'],info['Name'] or sku,int(t),s))
    rows.sort(key=lambda x:-x[3])
    grand=sum(r[3] for r in rows) or 1
    rows=rows[:10]

    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=9)
    c=ws.cell(2,1,f'{FULL} — TOP 10 TESTER PERFORMERS  (Jan–May 2026)')
    c.fill=fill(HDR); c.font=Font(name='Montserrat',bold=True,size=13,color='FFFFFFFF'); c.alignment=align()
    ws.row_dimensions[2].height=28
    for ci,h in enumerate(['Rank','Category','SKU Code','Product Name','Total Testers (T)',
                            'Total Sales (S)','Contrib% (T÷S)','% of Total Testers','Status'],1):
        H(ws,4,ci,h,HDR,s=9)
    ws.row_dimensions[4].height=28
    for rank,(sku,cat,name,t,s) in enumerate(rows,1):
        r=4+rank; contrib=t/s if s>0 else 0; pct=t/grand
        bg=GOLD if rank<=3 else (ALT_ROW if rank%2==0 else WHITE)
        D(ws,r,1,f'#{rank}',bg=bg,b=(rank<=3),s=10)
        D(ws,r,2,cat,bg=bg,h='left')
        D(ws,r,3,sku,bg=bg,fg=SKU_TEXT,b=True)
        D(ws,r,4,name,bg=bg,h='left')
        D(ws,r,5,t,bg=LIGHT_BLUE,b=True,s=10)
        D(ws,r,6,s,bg=LIGHT_GREEN,b=True,s=10)
        c7=ws.cell(r,7,contrib); c7.fill=fill(LIGHT_YELLOW); c7.font=font(s=9); c7.alignment=align(); c7.border=border(); c7.number_format='0.0%'
        c8=ws.cell(r,8,pct); c8.fill=fill(ALT_ROW); c8.font=font(s=9); c8.alignment=align(); c8.border=border(); c8.number_format='0.0%'
        st='★ Top Performer' if rank==1 else '▲ High Impact' if rank<=3 else ('◉ Monitor' if contrib>0.15 else '✓ Efficient')
        D(ws,r,9,st,bg=bg)
        ws.row_dimensions[r].height=20
    for ci,w in enumerate([6,20,11,34,16,14,14,18,16],1):
        ws.column_dimensions[get_column_letter(ci)].width=w

# ─── BUILD ───────────────────────────────────────────────────────────────────
# ─── ASL LEGACY SALES CODE REMAP ─────────────────────────────────────────────
# Jan–Apr 2026 ASL sales used old codes; testers/May use current codes.
# Confirmed: *_YAS001→YMK001, *_BAW001→BAW001, *_MAK001→MAK001, *_FUJ001/*_FJ001→FJ0001
# PROVISIONAL (geography certain, within-pair label per CLAUDE.md best-guess):
#   ASL_A009→YMK001, ASL_A011→BAS001 (both Abu Dhabi)
#   ASL_AL004→BAW001, ASL_AL007→MAK001 (both Al Ain)
ASL_SALES_REMAP = {
    'ASL_YAS001':'YMK001', 'ASL_BAW001':'BAW001', 'ASL_MAK001':'MAK001',
    'ASL_FUJ001':'FJ0001', 'ASL_FJ001':'FJ0001',
    'ASL_A009':'YMK001', 'ASL_A011':'BAS001',
    'ASL_AL004':'BAW001', 'ASL_AL007':'MAK001',
}

# ─── TESTER SOURCE = WAREHOUSE REPLENISHMENT DISPATCH (RptItemWiseStockTransfer) ─
#   Matches Vinayak's validated totals (White=24, More Of Oud SP0006=38, Safeena=9, Khaimah=15).
#   Store map: 0001→HO (Head Office), STR02→WH (FNF Warehouse); all else kept as-is.
#   Paper cards (EPT*/ASLT*) excluded. SKU codes are already bare FG codes in the file
#   (SP0006 = More Of Oud, O00006 = Oud Al Fakhamah — no mis-mapping).
def load_tester_dispatch():
    rep=pd.read_csv(os.path.join(OUT_DIR,'Replenishment_Jan_May_2026_FINAL.csv'))
    t=rep[rep.Item_Type=='Tester'].copy()
    t=t[~t.SKU_Code.astype(str).str.startswith(('EPT','ASLT'))]            # drop paper cards
    t=t[~t.Product_Name.str.contains('TESTER CARD|TESTER PAPER',case=False,na=False)]
    # TESTER CONSUMPTION = FG Warehouse → STORE only. Exclude Head Office (0001) and
    # FNF/RM Warehouse (STR02) — those are internal staging, not store consumption.
    t=t[~t.Store_Code.isin(['0001','STR02'])]
    t['store_code']=t.Store_Code
    MLBL={'Jan':'01','Feb':'02','Mar':'03','Apr':'04','May':'05','Jun':'06'}
    t['month_year']=t.Date.str.split('-').str[0].map(MLBL).radd('2026-')
    t=t[t.month_year.between('2026-01','2026-05')]
    t=t.rename(columns={'SKU_Code':'norm_sku','Brand':'division','Quantity':'testers'})
    t['norm_sku']=t['norm_sku'].replace(SKU_REMAP)
    return t.groupby(['norm_sku','store_code','division','month_year'],as_index=False)['testers'].sum()

# ─── SKU REMAP — code reconciliation (Vinayak-confirmed) ─────────────────────
# White Oud FG = SP0009 (boxed Set Box). Its testers + POS sales were mis-coded under
# the tester code SP0001 → roll both up to SP0009.
SKU_REMAP = {'SP0001':'SP0009'}

def build_report(brand):
    print(f'\n{"="*60}\nBuilding {brand} Tester Consumption Report (v3 — dispatch testers)...')
    catmap=load_category_map()
    T=load_tester_dispatch()
    S=pd.read_csv('/tmp/sales_fg.csv')
    S['sku_code']=S['sku_code'].replace(SKU_REMAP)   # SP0001→SP0009 White Oud
    S=S.groupby(['sku_code','store_code','month_year'],as_index=False)['qty_sold'].sum()

    grp=EPP_STORE_GROUPS if brand=='EPP' else ASL_STORE_GROUPS
    store_codes=[c for _,stores in grp for c,_ in stores]

    Tb=T[(T.division==brand)].copy()
    S=S.copy()
    if brand=='ASL':
        S['store_code']=S['store_code'].replace(ASL_SALES_REMAP)
        S=S.groupby(['sku_code','store_code','month_year'],as_index=False)['qty_sold'].sum()
    Sb=S[S.store_code.isin(store_codes)].copy()  # restrict sales to this brand's stores

    MONTHS=['2026-01','2026-02','2026-03','2026-04','2026-05']
    Q1=['2026-01','2026-02','2026-03']
    LBL={'2026-01':"Jan'26",'2026-02':"Feb'26",'2026-03':"Mar'26",'2026-04':"Apr'26",'2026-05':"May'26"}

    wb=Workbook(); wb.remove(wb.active)

    ws=wb.create_sheet("Top 10 Analysis"); ws.sheet_properties.tabColor='FFD700'
    build_top10(ws, brand, Tb, Sb, grp, catmap)
    print('  ✓ Top 10 Analysis')

    for m in MONTHS:
        ws=wb.create_sheet(LBL[m]); ws.sheet_properties.tabColor=('1F3864' if brand=='EPP' else '006064')
        n=build_matrix(ws, brand, Tb[Tb.month_year==m], Sb[Sb.month_year==m], grp, catmap)
        print(f'  ✓ {LBL[m]} — {n} SKUs')

    ws=wb.create_sheet("Q1'26 Matrix"); ws.sheet_properties.tabColor='388E3C'
    n=build_matrix(ws, brand, Tb[Tb.month_year.isin(Q1)], Sb[Sb.month_year.isin(Q1)], grp, catmap)
    print(f"  ✓ Q1'26 Matrix — {n} SKUs")

    path=os.path.join(OUT_DIR,f'Tester_Consumption_{brand}_Jan_May_2026_FINAL.xlsx')
    wb.save(path); print(f'  ✓ Saved: {path}')
    return path

if __name__=='__main__':
    e=build_report('EPP'); a=build_report('ASL')
    print('\n'+'='*60+'\nCOMPLETE:\n  EPP →',e,'\n  ASL →',a)
